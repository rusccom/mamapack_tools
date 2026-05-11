from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .settings import OUTPUT_PATH


def write_workbook(rows, audit, suppliers):
    wb = Workbook()
    ws = wb.active
    ws.title = "Spis z natury"
    write_spis_sheet(ws, rows, audit)
    write_summary_sheet(wb.create_sheet("Podsumowanie"), audit, suppliers)
    wb.save(OUTPUT_PATH)


def write_spis_sheet(ws, rows, audit):
    setup_page(ws)
    write_spis_header(ws, audit)
    write_headers(ws)
    for pos, row in enumerate(rows, start=1):
        write_spis_row(ws, pos, row)
    style_spis_table(ws, len(rows), len(spis_headers()))
    write_total_row(ws, len(rows))


def setup_page(ws):
    ws.freeze_panes = "A13"
    ws.sheet_view.showGridLines = False
    widths = [8, 24, 70, 12, 14, 14, 14, 18]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def write_spis_header(ws, audit):
    ws.merge_cells("A1:H1")
    ws["A1"] = "SMART TRADE RUSLAN YURCHENKO - Arkusz spisu z natury"
    ws["A1"].font = Font(bold=True, size=14)
    labels = [
        ("A3", "Arkusz spisu z natury nr", "1"),
        ("A4", "Przedmiot spisu", "Towary handlowe"),
        ("A5", "Spis z natury na dzień", audit["stock_day"]),
        ("A6", "Jednostka miary", "szt."),
        ("A7", "Wyłączono z wyceny", "Towary z dostaw TEMU"),
        ("A8", "Wygenerowano", audit["generated_at"]),
    ]
    for label_cell, label, value in labels:
        write_label(ws, label_cell, label, value)


def write_label(ws, cell, label, value):
    row = ws[cell].row
    ws[cell] = label
    ws.cell(row=row, column=3).value = value
    ws[cell].font = Font(bold=True)


def spis_headers():
    return ["Poz.", "Symbol indeksu lub kodu", "Nazwa składnika", "Jedn. miary",
            "Ilość", "Cena jednostkowa PLN", "Wartość PLN", "Uwagi"]


def write_headers(ws):
    for column, value in enumerate(spis_headers(), start=1):
        ws.cell(row=12, column=column).value = value


def write_spis_row(ws, pos, row):
    for column, value in enumerate(spis_values(pos, row), start=1):
        ws.cell(row=12 + pos, column=column).value = value


def spis_values(pos, row):
    note = f"BaseLinker ID {row['product_id']}; koszt: {row['source']}"
    return [pos, row["code"], row["name"], row["unit"], row["quantity"],
            row["unit_cost"], row["value"], note]


def style_spis_table(ws, row_count, col_count):
    header_row = 12
    last_row = header_row + row_count
    style_range(ws, header_row, last_row, col_count)
    table = Table(displayName="SpisZNatury", ref=f"A{header_row}:H{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)


def style_range(ws, header_row, last_row, col_count):
    for row in ws.iter_rows(min_row=header_row, max_row=last_row, max_col=col_count):
        style_cells(row, header_row)
    for row_number in range(header_row + 1, last_row + 1):
        apply_number_formats(ws, row_number)


def style_cells(row, header_row):
    for cell in row:
        cell.border = thin_border()
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if cell.row == header_row:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")


def apply_number_formats(ws, row_number):
    ws.cell(row_number, 5).number_format = "0.00"
    ws.cell(row_number, 6).number_format = "#,##0.00"
    ws.cell(row_number, 7).number_format = "#,##0.00"


def write_total_row(ws, row_count):
    total_row = 13 + row_count
    ws.cell(total_row, 6).value = "Razem"
    ws.cell(total_row, 7).value = f"=SUM(G13:G{total_row - 1})"
    ws.cell(total_row, 6).font = Font(bold=True)
    ws.cell(total_row, 7).font = Font(bold=True)
    ws.cell(total_row, 7).number_format = "#,##0.00"


def write_summary_sheet(ws, audit, suppliers):
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Podsumowanie spisu z natury"
    ws["A1"].font = Font(bold=True, size=14)
    rows = [["Pole", "Wartość"], *[[key, str(value)] for key, value in audit.items()]]
    for row in rows:
        ws.append(row)
    ws.append(["Dostawcy TEMU", ", ".join(temu_supplier_names(suppliers)) or "brak"])
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 72
    style_range(ws, 2, ws.max_row, 2)


def temu_supplier_names(suppliers):
    return sorted(name for name in suppliers.values() if "temu" in name.lower())


def thin_border():
    side = Side(style="thin", color="D9E2F3")
    return Border(left=side, right=side, top=side, bottom=side)
